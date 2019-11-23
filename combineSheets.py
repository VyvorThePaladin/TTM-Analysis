import openpyxl
import pandas as pd
import sys
# Load the two workbooks
# wb_orig = openpyxl.load_workbook('1-orig.xlsx')
# wb_adj = openpyxl.load_workbook('2-adj.xlsx')

wb_adj = pd.read_excel('2-adj.xlsx')
for index in range(0, len(wb_adj)):
    full_name = wb_adj.loc[index, 'Employee'].split()
    if len(full_name)<=2:
        wb_adj.loc[index, 'Employee'] = full_name[-1] + ', ' + full_name[0]
    elif len(full_name)>=3:
        length = len(full_name) 
        changed_name = full_name[-1] + ','
        for x in range(0, len(full_name)-1):
            changed_name = changed_name+' ' + full_name[x]
        wb_adj.loc[index, 'Employee'] = changed_name



wb_adj = wb_adj[['Employee UID', 'Employee', 'Adjustment Activity Code', 'Adjustment Decimal Hour']]
wb_adj['UID'] = wb_adj['Employee UID'] 
wb_adj['Employee Name'] = wb_adj['Employee']
wb_adj['Activity Code'] = wb_adj['Adjustment Activity Code']
wb_adj['hours_decimal'] = wb_adj['Adjustment Decimal Hour']
wb_adj = wb_adj[['UID', 'Employee Name', 'Activity Code', 'hours_decimal']]

wb_orig = pd.read_excel('1-orig.xlsx')
wb_orig = wb_orig[['UID', 'Employee Name', 'Activity Code', 'hours_decimal']]

combined = pd.concat([wb_adj, wb_orig]).reset_index(drop=True)
print(combined)
sys.exit()
wb_orig.loc[(wb_orig['Activity Code']=='1-Planning/PMER')|(wb_orig['Activity Code']=='1-Schematic Design'), 'Type'] = 'Project hours'
wb_orig = wb_orig.groupby(['Employee Name', 'Type'])


combined['Admin Hours']
combined['Project Hours']
combined['Sum'] = combined['Admin Hours']+combined['Project Hours']
combined['Admin %'] = combined['Admin Hours'] / combined['Sum'] * 100
combined['Admin %'] = combined['Admin %'].astype(int)


combined.to_excel("output.xlsx")



print(wb_adj)
# print(wb_orig.first())



# Creating new workbook and initializing titles
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active
sheet_new.title = 'TTM & Adj Combined'

# Get a handle on the original and adj TTM sheet
sheet_orig = wb_orig[wb_orig.sheetnames[0]]
sheet_adj = wb_adj[wb_adj.sheetnames[0]]

# Inserts Employee, UID, Activity Code, hours_decimal from orig into new sheet's columns
for i in range(1, sheet_orig.max_row+1):
    sheet_new['A' + str(i)] = sheet_orig['H' + str(i)].value
    sheet_new['B' + str(i)] = sheet_orig['G' + str(i)].value
    sheet_new['C' + str(i)] = sheet_orig['I' + str(i)].value
    sheet_new['D' + str(i)] = sheet_orig['M' + str(i)].value

#-----TO-DO-------#
# 1. Modify Employee name in adj_sheet and insert modified val in table
emp_list = []
for cellObj in sheet_adj['I']:
    emp_list.append(cellObj.value)

for name in emp_list:
    print(name)
wb_new.save('3-combinedRes.xlsx')

# Inserts Employee, UID, Activity Code, hours_decimal from orig into new sheet's columns
for j in range(2, sheet_adj.max_row+1):
    sheet_new['A' + str(sheet_orig.max_row+j-1)] = emp_list[j-1]    
    sheet_new['B' + str(sheet_orig.max_row+j-1)] = sheet_adj['B' + str(j)].value
    sheet_new['C' + str(sheet_orig.max_row+j-1)] = sheet_adj['D' + str(j)].value    
    sheet_new['D' + str(sheet_orig.max_row+j-1)] = sheet_adj['E' + str(j)].value
